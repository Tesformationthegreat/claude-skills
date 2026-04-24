"""
VERDICT template populator.

Fills the VERDICT scoring template with items and scores from the chat session,
applies the chosen weight preset, and saves a populated .xlsx.

Usage:
    python populate.py --items items.json --preset growth --output verdict-scored.xlsx

All original formulas, tier logic, and Red Flag checks in template.xlsx remain intact.
"""

import argparse
import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook


WEIGHT_PRESETS = {
    # Stage presets
    "startup":     {"V": 0.30, "E": 0.20, "R": 0.05, "D": 0.05, "I": 0.15, "C": 0.05, "T": 0.20},
    "growth":      {"V": 0.25, "E": 0.15, "R": 0.10, "D": 0.10, "I": 0.20, "C": 0.05, "T": 0.15},
    "enterprise":  {"V": 0.20, "E": 0.10, "R": 0.15, "D": 0.15, "I": 0.20, "C": 0.05, "T": 0.15},
    # Operating mode presets
    "survival":    {"V": 0.15, "E": 0.25, "R": 0.20, "D": 0.10, "I": 0.10, "C": 0.10, "T": 0.10},
    "innovation":  {"V": 0.30, "E": 0.10, "R": 0.05, "D": 0.05, "I": 0.20, "C": 0.05, "T": 0.25},
    "maintenance": {"V": 0.15, "E": 0.20, "R": 0.15, "D": 0.20, "I": 0.10, "C": 0.05, "T": 0.15},
}

# Column mapping on VERDICT Scoring tab
COL_ITEM = "B"
COL_TYPE = "C"
COL_KILL = "D"
COL_VALUE = "E"
COL_EFFORT = "F"
COL_RISK = "G"
COL_DEPS = "H"
COL_IMPACT = "I"
COL_CONFIDENCE = "J"
COL_TIME = "K"
COL_NOTES = "N"

# First data row on VERDICT Scoring tab (row 1 and 2 are headers)
DATA_START_ROW = 4


def load_items(path: str) -> list[dict]:
    with open(path, "r") as f:
        items = json.load(f)
    if not isinstance(items, list):
        raise ValueError("items.json must be a JSON array of item objects.")
    return items


def validate_item(item: dict, index: int) -> None:
    """Ensure required fields exist and scores are in range."""
    if "name" not in item or not item["name"]:
        raise ValueError(f"Item {index}: missing 'name' field.")

    kill = item.get("kill_criteria", "").upper()
    if kill not in ("PASS", "FAIL", ""):
        raise ValueError(f"Item {index} ({item['name']}): kill_criteria must be PASS, FAIL, or empty.")

    # If kill is FAIL, scores are not required
    if kill == "FAIL":
        return

    score_fields = ["value", "effort", "risk", "dependencies", "impact", "confidence", "time_criticality"]
    for field in score_fields:
        if field not in item:
            raise ValueError(f"Item {index} ({item['name']}): missing score '{field}'.")
        score = item[field]
        if not isinstance(score, (int, float)) or not (1 <= score <= 5):
            raise ValueError(f"Item {index} ({item['name']}): '{field}' must be a number 1 to 5, got {score}.")


def write_items(wb, items: list[dict]) -> None:
    """Write each item into the VERDICT Scoring tab."""
    if "VERDICT Scoring" not in wb.sheetnames:
        raise ValueError("Template is missing 'VERDICT Scoring' sheet. Check template.xlsx.")

    sheet = wb["VERDICT Scoring"]

    for i, item in enumerate(items):
        row = DATA_START_ROW + i
        sheet[f"{COL_ITEM}{row}"] = item["name"]
        sheet[f"{COL_TYPE}{row}"] = item.get("type", "")
        sheet[f"{COL_KILL}{row}"] = item.get("kill_criteria", "").upper()

        if item.get("kill_criteria", "").upper() != "FAIL":
            sheet[f"{COL_VALUE}{row}"] = item["value"]
            sheet[f"{COL_EFFORT}{row}"] = item["effort"]
            sheet[f"{COL_RISK}{row}"] = item["risk"]
            sheet[f"{COL_DEPS}{row}"] = item["dependencies"]
            sheet[f"{COL_IMPACT}{row}"] = item["impact"]
            sheet[f"{COL_CONFIDENCE}{row}"] = item["confidence"]
            sheet[f"{COL_TIME}{row}"] = item["time_criticality"]

        if item.get("notes"):
            sheet[f"{COL_NOTES}{row}"] = item["notes"]


def apply_preset(wb, preset: str) -> None:
    """Write the chosen preset weights into the Weights tab."""
    if preset not in WEIGHT_PRESETS:
        valid = ", ".join(WEIGHT_PRESETS.keys())
        raise ValueError(f"Unknown preset '{preset}'. Valid options: {valid}")

    if "Weights" not in wb.sheetnames:
        raise ValueError("Template is missing 'Weights' sheet. Check template.xlsx.")

    weights = WEIGHT_PRESETS[preset]
    sheet = wb["Weights"]

    # Weights live in column B, rows 3 through 9 (Value through Time Criticality)
    # Adjust these row numbers if the template layout changes.
    weight_rows = {
        "V": 3, "E": 4, "R": 5, "D": 6, "I": 7, "C": 8, "T": 9,
    }
    for dim, row in weight_rows.items():
        sheet[f"B{row}"] = weights[dim]


def main():
    parser = argparse.ArgumentParser(description="Populate the VERDICT template with scored items.")
    parser.add_argument("--items", required=True, help="Path to items.json")
    parser.add_argument("--preset", default="growth",
                        help="Weight preset: startup, growth, enterprise, survival, innovation, maintenance")
    parser.add_argument("--template", default="template.xlsx", help="Path to the VERDICT template")
    parser.add_argument("--output", default="verdict-scored.xlsx", help="Output path for populated file")
    args = parser.parse_args()

    template_path = Path(args.template)
    if not template_path.exists():
        print(f"Error: template not found at {template_path}", file=sys.stderr)
        sys.exit(1)

    items = load_items(args.items)
    for i, item in enumerate(items, start=1):
        validate_item(item, i)

    # Copy template so we never mutate the original
    shutil.copy(template_path, args.output)

    wb = load_workbook(args.output)
    apply_preset(wb, args.preset.lower())
    write_items(wb, items)
    wb.save(args.output)

    scored = sum(1 for it in items if it.get("kill_criteria", "").upper() != "FAIL")
    rejected = len(items) - scored
    print(f"Populated {args.output}: {scored} scored, {rejected} rejected, preset={args.preset}")


if __name__ == "__main__":
    main()
